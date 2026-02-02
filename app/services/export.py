"""
Export-Services für Dienstpläne (PDF und Excel)
"""
from datetime import date, timedelta
from calendar import monthrange
from io import BytesIO
from flask import make_response
from collections import defaultdict

from app.models import Mitarbeiter, Dienst, Dienstplan


class ExportService:
    """Service für PDF und Excel Export"""

    def export_pdf(self, jahr, monat):
        """
        Exportiert den Dienstplan als PDF

        Args:
            jahr: Jahr
            monat: Monat (1-12)

        Returns:
            Flask Response mit PDF
        """
        from weasyprint import HTML, CSS

        # Generate HTML
        html_content = self._generate_html(jahr, monat)

        # Convert to PDF
        pdf = HTML(string=html_content).write_pdf(
            stylesheets=[CSS(string=self._get_pdf_styles())]
        )

        # Create response
        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=dienstplan_{jahr}_{monat:02d}.pdf'

        return response

    def export_excel(self, jahr, monat):
        """
        Exportiert den Dienstplan als Excel

        Args:
            jahr: Jahr
            monat: Monat (1-12)

        Returns:
            Flask Response mit Excel-Datei
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        # Get data
        mitarbeiter = Mitarbeiter.query.filter_by(aktiv=True).order_by(Mitarbeiter.name).all()
        dienste = Dienst.query.all()
        _, num_days = monthrange(jahr, monat)

        start_datum = date(jahr, monat, 1)
        ende_datum = date(jahr, monat, num_days)

        dienstplaene = Dienstplan.query.filter(
            Dienstplan.datum >= start_datum,
            Dienstplan.datum <= ende_datum
        ).all()

        # Organize by date and employee
        plan_dict = {}
        for dp in dienstplaene:
            plan_dict[(dp.datum, dp.mitarbeiter_id)] = dp

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f'Dienstplan {monat:02d}/{jahr}'

        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        weekend_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')

        # Header row
        ws.cell(row=1, column=1, value='Mitarbeiter')
        ws.cell(row=1, column=1).font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).border = thin_border

        for tag in range(1, num_days + 1):
            d = date(jahr, monat, tag)
            weekday = ['Mo', 'Di', 'Mi', 'Do', 'Fr', 'Sa', 'So'][d.weekday()]
            cell = ws.cell(row=1, column=tag + 1, value=f'{weekday}\n{tag}')
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

            # Set column width
            ws.column_dimensions[get_column_letter(tag + 1)].width = 5

        # Set first column width
        ws.column_dimensions['A'].width = 20

        # Data rows
        for row_idx, ma in enumerate(mitarbeiter, start=2):
            ws.cell(row=row_idx, column=1, value=ma.name)
            ws.cell(row=row_idx, column=1).border = thin_border

            for tag in range(1, num_days + 1):
                d = date(jahr, monat, tag)
                cell = ws.cell(row=row_idx, column=tag + 1)
                cell.border = thin_border
                cell.alignment = center_align

                # Weekend highlighting
                if d.weekday() >= 5:
                    cell.fill = weekend_fill

                # Check for schedule
                dp = plan_dict.get((d, ma.id))
                if dp:
                    cell.value = dp.dienst.kurzname
                    # Color based on shift
                    color = dp.dienst.farbe.replace('#', '')
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    # Make text white for dark backgrounds
                    cell.font = Font(color='FFFFFF', bold=True)

        # Add legend
        legend_row = len(mitarbeiter) + 3
        ws.cell(row=legend_row, column=1, value='Legende:')
        ws.cell(row=legend_row, column=1).font = Font(bold=True)

        for idx, dienst in enumerate(dienste):
            row = legend_row + idx + 1
            ws.cell(row=row, column=1, value=dienst.kurzname)
            color = dienst.farbe.replace('#', '')
            ws.cell(row=row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws.cell(row=row, column=1).font = Font(color='FFFFFF', bold=True)
            ws.cell(row=row, column=2, value=f'{dienst.name} ({dienst.start_zeit.strftime("%H:%M")} - {dienst.ende_zeit.strftime("%H:%M")})')

        # Set row height for header
        ws.row_dimensions[1].height = 30

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=dienstplan_{jahr}_{monat:02d}.xlsx'

        return response

    def _generate_html(self, jahr, monat):
        """Generates HTML for PDF export"""
        # Get data
        mitarbeiter = Mitarbeiter.query.filter_by(aktiv=True).order_by(Mitarbeiter.name).all()
        dienste = Dienst.query.all()
        _, num_days = monthrange(jahr, monat)

        start_datum = date(jahr, monat, 1)
        ende_datum = date(jahr, monat, num_days)

        dienstplaene = Dienstplan.query.filter(
            Dienstplan.datum >= start_datum,
            Dienstplan.datum <= ende_datum
        ).all()

        plan_dict = {}
        for dp in dienstplaene:
            plan_dict[(dp.datum, dp.mitarbeiter_id)] = dp

        monat_name = ['Januar', 'Februar', 'März', 'April', 'Mai', 'Juni',
                      'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember'][monat - 1]

        html = f'''
<!DOCTYPE html>
<html lang="de">
<head>
    <meta charset="UTF-8">
    <title>Dienstplan {monat_name} {jahr}</title>
</head>
<body>
    <h1>Dienstplan {monat_name} {jahr}</h1>
    <table class="dienstplan">
        <thead>
            <tr>
                <th class="ma-name">Mitarbeiter</th>
'''

        for tag in range(1, num_days + 1):
            d = date(jahr, monat, tag)
            weekday = ['Mo', 'Di', 'Mi', 'Do', 'Fr', 'Sa', 'So'][d.weekday()]
            weekend_class = ' class="weekend"' if d.weekday() >= 5 else ''
            html += f'<th{weekend_class}>{weekday}<br>{tag}</th>\n'

        html += '''
            </tr>
        </thead>
        <tbody>
'''

        for ma in mitarbeiter:
            html += f'<tr><td class="ma-name">{ma.name}</td>\n'
            for tag in range(1, num_days + 1):
                d = date(jahr, monat, tag)
                weekend_class = ' class="weekend"' if d.weekday() >= 5 else ''
                dp = plan_dict.get((d, ma.id))
                if dp:
                    html += f'<td{weekend_class}><span class="dienst" style="background-color: {dp.dienst.farbe};">{dp.dienst.kurzname}</span></td>\n'
                else:
                    html += f'<td{weekend_class}>-</td>\n'
            html += '</tr>\n'

        html += '''
        </tbody>
    </table>

    <div class="legende">
        <h3>Legende</h3>
        <ul>
'''
        for dienst in dienste:
            html += f'<li><span class="dienst" style="background-color: {dienst.farbe};">{dienst.kurzname}</span> {dienst.name} ({dienst.start_zeit.strftime("%H:%M")} - {dienst.ende_zeit.strftime("%H:%M")})</li>\n'

        html += '''
        </ul>
    </div>

    <div class="footer">
        Erstellt am: ''' + date.today().strftime('%d.%m.%Y') + '''
    </div>
</body>
</html>
'''
        return html

    def _get_pdf_styles(self):
        """Returns CSS styles for PDF"""
        return '''
@page {
    size: A4 landscape;
    margin: 1cm;
}

body {
    font-family: Arial, sans-serif;
    font-size: 9pt;
}

h1 {
    font-size: 14pt;
    margin-bottom: 10px;
}

table.dienstplan {
    width: 100%;
    border-collapse: collapse;
    margin-bottom: 20px;
}

table.dienstplan th,
table.dienstplan td {
    border: 1px solid #ccc;
    padding: 2px 4px;
    text-align: center;
    vertical-align: middle;
}

table.dienstplan th {
    background-color: #0066CC;
    color: white;
    font-size: 8pt;
}

table.dienstplan th.ma-name,
table.dienstplan td.ma-name {
    text-align: left;
    min-width: 100px;
}

table.dienstplan .weekend {
    background-color: #f0f0f0;
}

.dienst {
    display: inline-block;
    padding: 2px 6px;
    border-radius: 3px;
    color: white;
    font-weight: bold;
    font-size: 8pt;
}

.legende {
    margin-top: 10px;
}

.legende h3 {
    font-size: 11pt;
    margin-bottom: 5px;
}

.legende ul {
    list-style: none;
    padding: 0;
    margin: 0;
}

.legende li {
    display: inline-block;
    margin-right: 15px;
    font-size: 8pt;
}

.footer {
    margin-top: 20px;
    font-size: 8pt;
    color: #666;
}
'''
