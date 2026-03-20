"""
XLSX-Import für Dienstplanwünsche der Mitarbeiter.

Liest das Format "Dienstplanwünsche MA" und mapped Präferenzen
auf die bestehenden Modelle (Dienst-Präferenzen, Regel-Ausnahmen,
Dienst-Einschränkungen).
"""
import re
import logging
from io import BytesIO
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Spalten-Mapping (0-basiert)
COL_NAME = 0        # A
COL_FRUEH = 1       # B
COL_KD = 2          # C
COL_TRIAGE = 3      # D
COL_SPAET = 4       # E
COL_NACHT = 5       # F
COL_NACHT_ANZAHL = 6   # G
COL_FRUEHE_NACHT = 7   # H
COL_SCHAUKELN = 8      # I
COL_GLEICHE_BLOCK = 9  # J
COL_FS_NACH_WE = 10    # K
COL_WE_DIENSTART = 11  # L
COL_ZNA_S11 = 12       # M
COL_LANGE_BLOECKE = 13  # N
COL_KURZE_BLOECKE = 14  # O
COL_PERSOENLICH = 15    # P


def parse_xlsx(file_or_path):
    """Parst die Dienstplanwünsche-XLSX und gibt strukturierte Daten zurück.

    Args:
        file_or_path: Dateipfad (str) oder BytesIO-Stream

    Returns:
        list[dict] mit einem Eintrag pro Mitarbeiter
    """
    if isinstance(file_or_path, (str, bytes)):
        wb = load_workbook(file_or_path, data_only=True)
    else:
        wb = load_workbook(file_or_path, data_only=True)

    ws = wb['Wünsche'] if 'Wünsche' in wb.sheetnames else wb[wb.sheetnames[0]]

    ergebnisse = []
    # Datenzeilen ab Zeile 5 (1-basiert), Zeile 3 = Header, Zeile 4 = Subheader
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        name = row[COL_NAME]
        if not name or not isinstance(name, str) or name.strip() == '':
            continue

        name = name.strip()

        eintrag = {
            'name': name,
            'dienste_gewuenscht': [],
            'nacht_config': _parse_nacht_anzahl(row[COL_NACHT_ANZAHL]),
            'fruehe_nacht': _parse_ja_nein(row[COL_FRUEHE_NACHT]),
            'schaukeln': _parse_ja_nein(row[COL_SCHAUKELN]),
            'gleiche_block': _parse_ja_nein(row[COL_GLEICHE_BLOCK]),
            'fs_nach_we': _parse_ja_nein(row[COL_FS_NACH_WE]),
            'we_dienstart': _normalize(row[COL_WE_DIENSTART]),
            'zna_s11': _normalize(row[COL_ZNA_S11]),
            'lange_bloecke': _is_checked(row[COL_LANGE_BLOECKE]),
            'kurze_bloecke': _is_checked(row[COL_KURZE_BLOECKE]),
            'persoenliche_wuensche': str(row[COL_PERSOENLICH]).strip() if row[COL_PERSOENLICH] else '',
        }

        # Dienste parsen (X = gewünscht, "egal"/"ausgeglichen" = alle ok)
        dienst_val = _normalize(row[COL_FRUEH])
        if dienst_val in ('x', 'egal', 'ausgeglichen'):
            eintrag['dienste_gewuenscht'].append('F')
        dienst_val = _normalize(row[COL_KD])
        if dienst_val in ('x', 'egal', 'ausgeglichen'):
            eintrag['dienste_gewuenscht'].append('KD')
        dienst_val = _normalize(row[COL_TRIAGE])
        if dienst_val in ('x', 'egal', 'ausgeglichen'):
            eintrag['dienste_gewuenscht'].append('T')
        dienst_val = _normalize(row[COL_SPAET])
        if dienst_val in ('x', 'egal', 'ausgeglichen'):
            eintrag['dienste_gewuenscht'].append('S')
        dienst_val = _normalize(row[COL_NACHT])
        if dienst_val in ('x', 'egal', 'ausgeglichen'):
            eintrag['dienste_gewuenscht'].append('N')

        # Wenn "egal" in Spalte B → alle Dienste
        if _normalize(row[COL_FRUEH]) in ('egal', 'ausgeglichen'):
            eintrag['dienste_gewuenscht'] = ['F', 'KD', 'T', 'S', 'N']

        # Einschränkungen aus Freitext parsen
        eintrag['einschraenkungen'] = _parse_freitext_einschraenkungen(
            eintrag['persoenliche_wuensche']
        )

        ergebnisse.append(eintrag)

    logger.info(f"XLSX geparst: {len(ergebnisse)} Mitarbeiter-Zeilen gefunden")
    return ergebnisse


def match_mitarbeiter(parsed_rows, mitarbeiter_liste):
    """Matcht XLSX-Namen auf bestehende Mitarbeiter.

    Returns:
        list[dict] mit 'xlsx_name', 'ma_id', 'ma_name', 'match_typ', 'daten'
    """
    ergebnisse = []

    for row in parsed_rows:
        xlsx_name = row['name']
        match = _find_best_match(xlsx_name, mitarbeiter_liste)

        ergebnisse.append({
            'xlsx_name': xlsx_name,
            'ma_id': match['id'] if match else None,
            'ma_name': match['name'] if match else None,
            'match_typ': match['typ'] if match else 'keine',
            'daten': row,
        })

    return ergebnisse


def _find_best_match(xlsx_name, mitarbeiter_liste):
    """Findet den besten Mitarbeiter-Match für einen XLSX-Namen."""
    name_lower = xlsx_name.lower().strip().rstrip('.')

    # 1. Exakter Vollname-Match
    for m in mitarbeiter_liste:
        if m.name.lower() == name_lower:
            return {'id': m.id, 'name': m.name, 'typ': 'exakt'}

    # 2. Mehrteilige XLSX-Namen: exakter Match auf Vor+Nachname
    xlsx_teile = name_lower.split()
    if len(xlsx_teile) >= 2:
        for m in mitarbeiter_liste:
            if all(teil in m.name.lower() for teil in xlsx_teile):
                return {'id': m.id, 'name': m.name, 'typ': 'exakt'}

    # 3. Teilstring-Match (Vorname/Spitzname)
    kandidaten = []
    for m in mitarbeiter_liste:
        db_name_lower = m.name.lower()
        db_teile = db_name_lower.split()

        best_score = 0
        for teil in xlsx_teile:
            teil_clean = teil.rstrip('.')
            if len(teil_clean) < 3:
                continue
            for db_teil in db_teile:
                # Exakter Wort-Match (höchster Score)
                if teil_clean == db_teil:
                    best_score = max(best_score, len(teil_clean) * 2)
                # Präfix-Match
                elif db_teil.startswith(teil_clean) or teil_clean.startswith(db_teil):
                    overlap = min(len(teil_clean), len(db_teil))
                    best_score = max(best_score, overlap)
                # Substring-Match
                elif teil_clean in db_teil or db_teil in teil_clean:
                    overlap = min(len(teil_clean), len(db_teil))
                    best_score = max(best_score, overlap)
                # Fuzzy: gleiche Anfangsbuchstaben + ähnliche Länge (Tippfehler)
                elif (len(teil_clean) >= 5 and len(db_teil) >= 5 and
                      teil_clean[:4] == db_teil[:4] and
                      abs(len(teil_clean) - len(db_teil)) <= 2):
                    best_score = max(best_score, len(teil_clean) - 1)

        if best_score > 0:
            kandidaten.append({'id': m.id, 'name': m.name, 'typ': 'teilmatch',
                               'score': best_score})

    if kandidaten:
        kandidaten.sort(key=lambda x: x['score'], reverse=True)
        best = kandidaten[0]
        # Mehrdeutigkeit prüfen
        same_score = [k for k in kandidaten if k['score'] == best['score']]
        if len(same_score) > 1:
            best['typ'] = 'mehrdeutig'
        return best

    return None


def _parse_nacht_anzahl(val):
    """Parst die Nacht-Anzahl-Spalte in strukturierte Daten.

    Beispiele: "2x3", "viele", "keine", "1x4", "2x2 / 2x3", "max 2x3", "14 Tage"
    """
    if not val:
        return None

    text = str(val).strip().lower()

    if text in ('keine', 'kein', '0'):
        return {'typ': 'keine', 'max_naechte': 0}

    if text in ('viele', 'viel'):
        return {'typ': 'viele'}

    # "14 Tage" o.ä.
    tage_match = re.match(r'(\d+)\s*tage?', text)
    if tage_match:
        return {'typ': 'tage', 'tage': int(tage_match.group(1))}

    # "max 2x3"
    max_match = re.match(r'max\s*(\d+)x(\d+)', text)
    if max_match:
        bloecke = int(max_match.group(1))
        laenge = int(max_match.group(2))
        return {'typ': 'block', 'max_bloecke': bloecke, 'max_laenge': laenge,
                'max_naechte': bloecke * laenge}

    # "2x3" oder "2x2 / 2x3" oder "2x2/2x3"
    block_matches = re.findall(r'(\d+)x(\d+)', text)
    if block_matches:
        total_naechte = 0
        min_laenge = 99
        max_laenge = 0
        total_bloecke = 0
        for b, l in block_matches:
            bloecke = int(b)
            laenge = int(l)
            total_naechte += bloecke * laenge
            total_bloecke += bloecke
            min_laenge = min(min_laenge, laenge)
            max_laenge = max(max_laenge, laenge)
        return {
            'typ': 'block',
            'total_bloecke': total_bloecke,
            'min_laenge': min_laenge,
            'max_laenge': max_laenge,
            'max_naechte': total_naechte,
        }

    return {'typ': 'unbekannt', 'roh': text}


def _parse_freitext_einschraenkungen(text):
    """Versucht parsbare Einschränkungen aus dem Freitext zu extrahieren.

    Returns:
        list[dict] mit 'tag', 'typ', 'dienst', 'beschreibung'
    """
    if not text:
        return []

    ergebnisse = []
    text_lower = text.lower()

    # Wochentags-Patterns
    tag_mapping = {
        'montag': 'MONTAG', 'mo': 'MONTAG',
        'dienstag': 'DIENSTAG', 'di': 'DIENSTAG',
        'mittwoch': 'MITTWOCH', 'mi': 'MITTWOCH',
        'donnerstag': 'DONNERSTAG', 'do': 'DONNERSTAG',
        'freitag': 'FREITAG', 'fr': 'FREITAG',
        'samstag': 'SAMSTAG', 'sa': 'SAMSTAG',
        'sonntag': 'SONNTAG', 'so': 'SONNTAG',
    }

    dienst_mapping = {
        'früh': 'Frühdienst', 'fd': 'Frühdienst', 'f': 'Frühdienst',
        'spät': 'Spätdienst', 'sms': 'Spätdienst', 'sd': 'Spätdienst',
        'nacht': 'Nachtdienst', 'nd': 'Nachtdienst',
        'kd': 'Kerndienst', 'kerndienst': 'Kerndienst',
        'triage': 'Triage',
    }

    # Pattern: "[Wochentag]s kein/keine [Dienst]" oder "[Wochentag]s frei"
    for tag_word, tag_typ in tag_mapping.items():
        # "freitags frei" / "freitags keine dienste"
        pattern_frei = rf'{tag_word}s?\s+(?:kein(?:e)?|frei)'
        if re.search(pattern_frei, text_lower):
            ergebnisse.append({
                'tag': tag_typ,
                'typ': 'frei',
                'dienst': None,
                'beschreibung': f'{tag_word.capitalize()} frei'
            })

        # "freitags kein Spät" / "Mo/Di/Mi kein Früh"
        for dienst_word, dienst_name in dienst_mapping.items():
            pattern = rf'{tag_word}s?\s+kein(?:e|en)?\s+{dienst_word}'
            if re.search(pattern, text_lower):
                ergebnisse.append({
                    'tag': tag_typ,
                    'typ': 'kein_dienst',
                    'dienst': dienst_name,
                    'beschreibung': f'{tag_word.capitalize()} kein {dienst_name}'
                })

    # Kompakt-Pattern: "Mo/Di/Mi kein Früh"
    kompakt_match = re.search(r'((?:mo|di|mi|do|fr|sa|so)(?:/(?:mo|di|mi|do|fr|sa|so))+)\s+kein(?:e|en)?\s+(\w+)', text_lower)
    if kompakt_match:
        tage_str = kompakt_match.group(1)
        dienst_str = kompakt_match.group(2)
        dienst_name = dienst_mapping.get(dienst_str, dienst_str)
        for tag_kurz in tage_str.split('/'):
            tag_typ = tag_mapping.get(tag_kurz.strip())
            if tag_typ:
                # Prüfen ob schon vorhanden
                exists = any(e['tag'] == tag_typ and e['dienst'] == dienst_name for e in ergebnisse)
                if not exists:
                    ergebnisse.append({
                        'tag': tag_typ,
                        'typ': 'kein_dienst',
                        'dienst': dienst_name,
                        'beschreibung': f'{tag_kurz.upper()} kein {dienst_name}'
                    })

    # "nur 1 WE im Monat"
    we_match = re.search(r'(?:nur|max)\s+(\d+)\s+we', text_lower)
    if we_match:
        ergebnisse.append({
            'tag': 'WOCHENENDE',
            'typ': 'max_we',
            'wert': int(we_match.group(1)),
            'beschreibung': f'Max {we_match.group(1)} WE/Monat'
        })

    return ergebnisse


def _normalize(val):
    """Normalisiert einen Zellwert zu lowercase String."""
    if val is None:
        return ''
    return str(val).strip().lower()


def _is_checked(val):
    """Prüft ob eine Zelle als 'angekreuzt' gilt."""
    return _normalize(val) == 'x'


def _parse_ja_nein(val):
    """Parst ja/nein/egal-Werte."""
    text = _normalize(val)
    if text in ('ja', 'eher ja'):
        return 'ja'
    elif text in ('nein', 'eher nein'):
        return 'nein'
    elif text in ('egal', 'n.r.', 'n.r', ''):
        return 'egal'
    return text


def importiere_praeferenzen(matched_data, dienst_map, db_session):
    """Importiert die geparsten Präferenzen in die Datenbank.

    Args:
        matched_data: Ergebnis von match_mitarbeiter() (nur Einträge mit ma_id)
        dienst_map: Dict {kurzname -> Dienst-Objekt} z.B. {'F': Dienst(Frühdienst)}
        db_session: SQLAlchemy Session

    Returns:
        dict mit Import-Statistiken
    """
    from app.models import (
        Mitarbeiter, MitarbeiterDienstPraeferenz,
        MitarbeiterDienstEinschraenkung, TagTyp
    )

    stats = {
        'praeferenzen_erstellt': 0,
        'regel_ausnahmen_gesetzt': 0,
        'einschraenkungen_erstellt': 0,
        'uebersprungen': 0,
        'fehler': [],
        'hinweise': [],
    }

    for eintrag in matched_data:
        ma_id = eintrag.get('ma_id')
        if not ma_id:
            stats['uebersprungen'] += 1
            continue

        ma = Mitarbeiter.query.get(ma_id)
        if not ma:
            stats['fehler'].append(f"{eintrag['xlsx_name']}: MA ID {ma_id} nicht gefunden")
            continue

        daten = eintrag['daten']

        # 1. Bestehende Präferenzen löschen und neu anlegen
        MitarbeiterDienstPraeferenz.query.filter_by(mitarbeiter_id=ma_id).delete()

        for kuerzel in daten['dienste_gewuenscht']:
            dienst = dienst_map.get(kuerzel)
            if dienst:
                pref = MitarbeiterDienstPraeferenz(
                    mitarbeiter_id=ma_id,
                    dienst_id=dienst.id,
                    min_pro_monat=0,
                    max_pro_monat=None,
                )
                db_session.add(pref)
                stats['praeferenzen_erstellt'] += 1

        # 2. Nacht-Konfiguration → regel_ausnahmen
        nacht = daten.get('nacht_config')
        if nacht:
            ausnahmen = ma.regel_ausnahmen.copy() if ma.regel_ausnahmen else {}

            if nacht['typ'] == 'keine':
                ausnahmen['MAX_NAECHTE_MONAT'] = 0
                ausnahmen['MIN_NAECHTE_MONAT'] = 0
                stats['regel_ausnahmen_gesetzt'] += 1
            elif nacht['typ'] == 'viele':
                # Keine Beschränkung → ggf. vorhandene Limits entfernen
                ausnahmen.pop('MAX_NAECHTE_MONAT', None)
                stats['regel_ausnahmen_gesetzt'] += 1
            elif nacht['typ'] == 'block' and 'max_naechte' in nacht:
                ausnahmen['MAX_NAECHTE_MONAT'] = nacht['max_naechte']
                stats['regel_ausnahmen_gesetzt'] += 1

            ma.regel_ausnahmen = ausnahmen

        # 3. Einschränkungen aus Freitext
        for einschr in daten.get('einschraenkungen', []):
            tag_typ_str = einschr.get('tag')
            if not tag_typ_str:
                continue

            try:
                tag_typ = TagTyp(tag_typ_str)
            except ValueError:
                continue

            if einschr['typ'] == 'frei':
                # "Freitags frei" → kann nicht direkt als Einschränkung modelliert werden
                # (müsste Wunsch pro Datum sein). Notiz speichern.
                stats['hinweise'].append(
                    f"{ma.name}: '{einschr['beschreibung']}' – bitte manuell als Wunsch eintragen"
                )
                continue

            if einschr['typ'] == 'kein_dienst' and einschr.get('dienst'):
                # Finde den Dienst
                target_dienst = None
                for d in dienst_map.values():
                    if d.name == einschr['dienst']:
                        target_dienst = d
                        break

                if not target_dienst:
                    stats['hinweise'].append(
                        f"{ma.name}: Dienst '{einschr['dienst']}' nicht gefunden für Einschränkung"
                    )
                    continue

                # "Mo kein Früh" = an Mo darf NICHT Früh → invertiert:
                # Wir können nur "nur Dienst X erlaubt" modellieren.
                # Stattdessen als Hinweis speichern.
                stats['hinweise'].append(
                    f"{ma.name}: '{einschr['beschreibung']}' – "
                    f"Einschränkungsmodell unterstützt nur 'nur Dienst X erlaubt', nicht 'kein Dienst Y'. "
                    f"Bitte manuell prüfen."
                )

            if einschr['typ'] == 'max_we' and 'wert' in einschr:
                ausnahmen = ma.regel_ausnahmen.copy() if ma.regel_ausnahmen else {}
                ausnahmen['WOCHENENDE_ROTATION'] = einschr['wert']
                ma.regel_ausnahmen = ausnahmen
                stats['regel_ausnahmen_gesetzt'] += 1

        # 4. Persönliche Wünsche als Hinweis
        if daten.get('persoenliche_wuensche'):
            stats['hinweise'].append(
                f"{ma.name}: \"{daten['persoenliche_wuensche']}\""
            )

    try:
        db_session.commit()
        logger.info(f"Import abgeschlossen: {stats}")
    except Exception as e:
        db_session.rollback()
        stats['fehler'].append(f"Datenbankfehler: {str(e)}")
        logger.error(f"Import-Fehler: {e}")

    return stats
